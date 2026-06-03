#!/usr/bin/env python3
# 거래내역 엑셀 → price_history.json 생성
# 키별로 판매(출고)/매입(입고)/견적(견적) 세 가지 이력을 분리 저장한다.
# 출력 포맷(키 1개): [basis, n, min, max, vol, sales_samples, buy_samples, quo_samples]
#   basis: 0=출고기준, 1=견적기준(출고 없을 때 가격 산정에 사용)
#   n/min/max/vol: 가격 산정 스트림(출고 우선, 없으면 견적)의 건수/최저/최고/변동성
#   *_samples: [date, price, cust, qty, code] 의 리스트(최신순)
import openpyxl, re, json, statistics, sys
from collections import defaultdict
from datetime import datetime, timedelta
import bisect

src = sys.argv[1] if len(sys.argv) > 1 else "/root/.claude/uploads/bbe1baef-4c85-4af7-ba35-e24511b32496/826173da-____________20260602_100018.xlsx"
SALE_CAP = int(sys.argv[2]) if len(sys.argv) > 2 else 60   # 판매이력 최대 표시 건수
BUY_CAP  = int(sys.argv[3]) if len(sys.argv) > 3 else 40   # 매입이력 최대 표시 건수
QUO_CAP  = int(sys.argv[4]) if len(sys.argv) > 4 else 40   # 견적이력 최대 표시 건수

wb = openpyxl.load_workbook(src, read_only=True, data_only=True); ws = wb["데이터"]
MAT = {"VITON":"FKM","바이톤":"FKM","FKM":"FKM","EPDM":"EPDM","SILICONE":"SILICON","SILICON":"SILICON","실리콘":"SILICON","우레탄":"PU","URETHANE":"PU","PTFE":"PTFE","테프론":"PTFE","TEFLON":"PTFE","POM":"POM","NYLON":"NYLON","나일론":"NYLON","페놀":"PHENOLIC","HNBR":"HNBR","NBR":"NBR","니트릴":"NBR"}
def mat_kw(s):
    u=s.upper()
    for k,v in MAT.items():
        if k in u: return v
    return ""
def infer_mat(code,name):
    s=(str(name)+" "+str(code))
    m=mat_kw(s)
    if m: return m
    c=str(code).upper().replace(" ","")
    if re.search(r'(FT|F|V)$',c): return "FKM"
    if c.endswith("Q"): return "SILICON"
    U=s.upper()
    if any(x in U for x in ("WEAR","BACK","GASKET","PTFE","NYLON","페놀","BUSH","CARBON","BRONZE")): return ""
    return "NBR"
def norm(x):
    f=round(float(x),2); return ("%f"%f).rstrip("0").rstrip(".")
CODE_RE=re.compile(r'(AS568-?\d+|P\d+(?:\.\d+)?[A-Z]?|G\d+[A-Z]?|S\d+[A-Z]?|AN\s?\d+|K\d+)')
DIM3=re.compile(r'(\d+\.?\d*)\s*[*xX]\s*(\d+\.?\d*)\s*[*xX]\s*(\d+\.?\d*)')
DIM2=re.compile(r'(\d+\.?\d*)\s*[*xX]\s*(\d+\.?\d*)')
TYPE_RE=re.compile(r'\b(TCN|TCV|TC|SC|SB|TB|VF|VH|DSI|SKY|VA|VS)\b')
def keys_for(code,name,mat):
    ks=set(); text=(str(code)+" "+str(name)).upper().replace("*","x")
    cm=CODE_RE.search(text)
    if cm:
        c=re.sub(r'[\s-]','',cm.group(1))
        if re.match(r'^(AN|P|G|S|K)\d',c): c=re.sub(r'(FT|F|Q|J|N|E|S|V|T)$','',c)
        ks.add("C:"+c)
        if mat: ks.add("C:"+c+"|"+mat)
    tm=TYPE_RE.search(text); d3=DIM3.search(text)
    if tm and d3:
        t=tm.group(1); a,b,c=[norm(x) for x in d3.groups()]; ks.add("D:%s:%sx%sx%s"%(t,a,b,c))
        if mat: ks.add("D:%s:%sx%sx%s|%s"%(t,a,b,c,mat))
    if d3 and not tm:
        a,b,c=[norm(x) for x in d3.groups()]; ks.add("D::%sx%sx%s"%(a,b,c))
        if mat: ks.add("D::%sx%sx%s|%s"%(a,b,c,mat))
    d2=DIM2.search(text)
    if d2 and not d3:
        a,b=[norm(x) for x in d2.groups()]; ks.add("D::%sx%s"%(a,b))
        if mat: ks.add("D::%sx%s|%s"%(a,b,mat))
    return ks
EXCLUDE=("금형","등분 견적","택배","운임","DC","할인","운반","샘플")
def ncust(c): return str(c or "").strip()
def ncode(c): return re.sub(r'[\s\-./]','',str(c or "").upper())

# 1차 패스: 견적→거래 성사 판정을 위해 (거래처,품목코드)별 출고일자 인덱스 구축
SELL_IDX=defaultdict(list); MAXD=None
rows=ws.iter_rows(values_only=True); next(rows)
for r in rows:
    code,name,cust,date,gb,qty,price,amt=(list(r)+[None]*8)[:8]
    if not isinstance(date,datetime): continue
    if MAXD is None or date>MAXD: MAXD=date
    if str(gb)=="출고" and code: SELL_IDX[(ncust(cust),ncode(code))].append(date)
for k in SELL_IDX: SELL_IDX[k].sort()
RECENT_DAYS=60  # 견적일이 데이터 최신일 기준 이 기간 이내면 '대기중'(아직 판단 보류)
def conv_flag(cust,code,qd):
    # 2=성사(견적 후 1년 내 같은 거래처에 같은 코드 출고), 1=대기중(최근 견적), 0=미성사
    if not isinstance(qd,datetime): return 0
    ds=SELL_IDX.get((ncust(cust),ncode(code)))
    if ds:
        lo=qd-timedelta(days=7); hi=qd+timedelta(days=365)
        i=bisect.bisect_left(ds,lo)
        if i<len(ds) and ds[i]<=hi: return 2
    if MAXD and (MAXD-qd).days<=RECENT_DAYS: return 1
    return 0

won=defaultdict(list); buy=defaultdict(list); quo=defaultdict(list)
rows=ws.iter_rows(values_only=True); next(rows)
for r in rows:
    code,name,cust,date,gb,qty,price,amt=(list(r)+[None]*8)[:8]
    if not code or not name: continue
    if any(x in str(name) for x in EXCLUDE): continue
    try: p=float(price)
    except: p=0
    if p<=0 or p>2000000: continue
    try: q=int(qty) if qty is not None else 0
    except: q=0
    mat=infer_mat(code,name)
    rec=(str(date)[:10],round(p),str(cust or "").strip()[:16],q,str(code).strip()[:16])
    g=str(gb)
    if g=="출고":   dst=won
    elif g=="입고": dst=buy
    elif g=="견적": dst=quo; rec=rec+(conv_flag(cust,code,date),) # 6번째: 거래성사 플래그
    else: continue
    for k in keys_for(code,name,mat): dst[k].append(rec)
def vol_of(pr):
    ps=pr[:12]
    if len(ps)<2: return 0
    m=sum(ps)/len(ps)
    return 1 if m>0 and (statistics.pstdev(ps)/m>0.25 or max(ps)/max(min(ps),1)>1.8) else 0
def srt(lst,cap):
    return [list(x) for x in sorted(lst,key=lambda x:x[0],reverse=True)[:cap]]
out={}
for k in set(won)|set(buy)|set(quo):
    w=sorted(won.get(k,[]),key=lambda x:x[0],reverse=True)
    b=sorted(buy.get(k,[]),key=lambda x:x[0],reverse=True)
    q=sorted(quo.get(k,[]),key=lambda x:x[0],reverse=True)
    basis,srcv=(0,w) if w else (1,q)
    if not srcv:  # 출고·견적 모두 없고 입고만 있으면 매입가만 보여줌(가격산정 불가)
        prices=[]; n=0; mn=0; mx=0; vol=0
    else:
        prices=[x[1] for x in srcv]; n=len(prices); mn=round(min(prices)); mx=round(max(prices)); vol=vol_of(prices)
    out[k]=[basis,n,mn,mx,vol,
            [[x[0],x[1],x[2],x[3],x[4]] for x in w[:SALE_CAP]],
            [[x[0],x[1],x[2],x[3],x[4]] for x in b[:BUY_CAP]],
            [[x[0],x[1],x[2],x[3],x[4],x[5]] for x in q[:QUO_CAP]]]  # 견적: 6번째=성사플래그
j=json.dumps(out,ensure_ascii=False,separators=(",",":"))
open("price_history.json","w",encoding="utf-8").write(j)
print("키:",len(out),"| 크기:",round(len(j.encode())/1024/1024,2),"MB","| caps",SALE_CAP,BUY_CAP,QUO_CAP)
for t in ["C:P50|NBR","C:P50|FKM","C:P50","C:P7"]:
    v=out.get(t)
    if v: print(" ",t,"-> n=%d 판매%d 매입%d 견적%d 최근 %s"%(v[1],len(v[5]),len(v[6]),len(v[7]),(str(v[5][0]) if v[5] else (str(v[7][0]) if v[7] else '-'))))
    else: print(" ",t,"없음")
